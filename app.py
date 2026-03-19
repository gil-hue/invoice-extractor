import streamlit as st
import pandas as pd
import io
import os
import re
import yaml
import json
import base64
import requests
import bcrypt
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit_authenticator as stauth
from yaml.loader import SafeLoader
import anthropic

LOGS_FILE = "logs.json"

# ── Column definitions ─────────────────────────────────────────
COL_ORDER = [
    "source_file", "invoice_number", "date", "supplier_name", "supplier_vat",
    "customer_name", "customer_vat", "document_type", "payment_method",
    "amount_before_vat", "vat_amount", "total_amount", "notes",
]
COL_LABELS = {
    "source_file":       "📄 שם קובץ",
    "invoice_number":    "מספר חשבונית",
    "date":              "תאריך",
    "supplier_name":     "שם ספק",
    "supplier_vat":      'ע"מ ספק',
    "customer_name":     "שם לקוח",
    "customer_vat":      'ע"מ לקוח',
    "document_type":     "סוג מסמך",
    "payment_method":    "אמצעי תשלום",
    "amount_before_vat": 'סכום לפני מע"מ',
    "vat_amount":        'מע"מ',
    "total_amount":      "סכום כולל",
    "notes":             "הערות",
}

EXTRACT_PROMPT = """נתח את המסמך/החשבונית הזו והחזר JSON עם השדות הבאים בדיוק:
{
  "invoice_number": "מספר החשבונית",
  "date": "תאריך בפורמט DD/MM/YYYY",
  "supplier_name": "שם הספק/העסק המוציא",
  "supplier_vat": "מספר ע.מ./ח.פ. של הספק",
  "customer_name": "שם הלקוח",
  "customer_vat": "מספר ע.מ./ח.פ. של הלקוח",
  "amount_before_vat": "סכום לפני מע\"מ כמספר עשרוני (לדוגמה: 1000.00)",
  "vat_amount": "סכום מע\"מ כמספר עשרוני (לדוגמה: 170.00)",
  "total_amount": "סכום כולל כמספר עשרוני (לדוגמה: 1170.00)",
  "document_type": "אחד מ: חשבונית מס / קבלה / חשבונית מס קבלה / תעודת משלוח / הצעת מחיר / אחר",
  "payment_method": "אחד מ: מזומן / אשראי / העברה בנקאית / צ'ק / לא צוין",
  "notes": "תיאור קצר של השירות/המוצר ופרטים נוספים רלוונטיים"
}
החזר JSON בלבד, ללא טקסט לפני או אחרי. אם ערך לא נמצא — השתמש במחרוזת ריקה ""."""

# ── Page config ────────────────────────────────────────────────
st.set_page_config(page_title="חשבוניות חכמות", page_icon="🧾", layout="wide")

# ── CSS ────────────────────────────────────────────────────────
st.markdown("""
<style>
div[data-testid="stForm"] {
    max-width: 420px; margin: 2rem auto; background: white;
    padding: 2rem 2.5rem 1.5rem; border-radius: 16px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.12);
}
div[data-testid="stTextInput"] input {
    font-size: 0.88rem !important; padding: 0.35rem 0.6rem !important; height: 36px !important;
}
div[data-testid="stTextInput"] label { font-size: 0.85rem !important; }
div[data-testid="stForm"] button[kind="primaryFormSubmit"] {
    width: 100%; background: linear-gradient(135deg, #1A4731, #2E7D52) !important;
    color: white !important; border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; margin-top: 0.5rem;
}
.main { background-color: #f8fbf9; }
.block-container { padding-top: 1.5rem; }
.title-bar {
    background: linear-gradient(135deg, #1A4731 0%, #2E7D52 100%);
    padding: 1.2rem 2rem; border-radius: 12px; margin-bottom: 1.5rem; color: white;
}
.title-bar h1 { color: white; margin: 0; font-size: 1.8rem; }
.title-bar p  { color: #b8e8c8; margin: 0.2rem 0 0 0; font-size: 0.95rem; }
.metric-card {
    background: white; border-radius: 10px; padding: 1.2rem;
    text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    border-left: 4px solid #2E7D52;
}
.metric-card .num { font-size: 2rem; font-weight: 700; color: #1A4731; }
.metric-card .lbl { font-size: 0.85rem; color: #666; margin-top: 0.2rem; }
.stButton > button {
    background: linear-gradient(135deg, #1A4731, #2E7D52);
    color: white; border: none; border-radius: 8px;
    padding: 0.6rem 2rem; font-size: 1rem; font-weight: 600; width: 100%;
}
.stButton > button:hover { opacity: 0.9; }
.stButton > button[kind="secondary"] {
    background: #f0f7f3 !important; color: #1A4731 !important;
    border: 1px solid #b8dbc8 !important; font-size: 0.88rem !important;
    font-weight: 500 !important; padding: 0.3rem 0.5rem !important;
}
.stButton > button[kind="secondary"]:hover {
    background: #d4edde !important; border-color: #2E7D52 !important;
}
.success-box {
    background: #e8f5e9; border-left: 4px solid #43a047;
    padding: 1rem 1.2rem; border-radius: 8px; color: #2e7d32; font-weight: 600;
}
.section-header {
    font-size: 1.1rem; font-weight: 700; color: #1A4731;
    border-bottom: 2px solid #2E7D52; padding-bottom: 0.4rem; margin-bottom: 1rem;
}
.login-title { text-align: center; color: #1A4731; font-size: 1.6rem; font-weight: 700; margin-bottom: 0.2rem; }
.login-sub   { text-align: center; color: #888; font-size: 0.9rem; margin-bottom: 1.5rem; }
</style>
""", unsafe_allow_html=True)

# ── Load config ────────────────────────────────────────────────
with open("config.yaml") as f:
    config = yaml.load(f, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config["credentials"],
    config["cookie"]["name"],
    config["cookie"]["key"],
    config["cookie"]["expiry_days"],
)

# ── Log helpers ────────────────────────────────────────────────
def read_logs():
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO", "gil-hue/invoice-extractor")
        if token:
            url = f"https://api.github.com/repos/{repo}/contents/logs.json"
            r   = requests.get(url, headers={"Authorization": f"token {token}"})
            if r.status_code == 200:
                return json.loads(base64.b64decode(r.json()["content"]).decode())
    except Exception:
        pass
    try:
        if os.path.exists(LOGS_FILE):
            with open(LOGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list) and data:
                    return data
    except Exception:
        pass
    return st.session_state.get("session_logs", [])


def write_log(action, details, user=None):
    entry = {
        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "user":      user or st.session_state.get("username", "—"),
        "action":    action,
        "details":   details,
    }
    if "session_logs" not in st.session_state:
        st.session_state["session_logs"] = []
    st.session_state["session_logs"].insert(0, entry)
    try:
        logs = []
        if os.path.exists(LOGS_FILE):
            with open(LOGS_FILE, "r", encoding="utf-8") as f:
                logs = json.load(f)
        if not isinstance(logs, list):
            logs = []
        logs.insert(0, entry)
        with open(LOGS_FILE, "w", encoding="utf-8") as f:
            json.dump(logs[:500], f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO", "gil-hue/invoice-extractor")
        if not token:
            return
        url     = f"https://api.github.com/repos/{repo}/contents/logs.json"
        headers = {"Authorization": f"token {token}"}
        r       = requests.get(url, headers=headers)
        if r.status_code != 200:
            return
        sha  = r.json()["sha"]
        logs = json.loads(base64.b64decode(r.json()["content"]).decode())
        logs.insert(0, entry)
        content = base64.b64encode(
            json.dumps(logs[:500], ensure_ascii=False, indent=2).encode()
        ).decode()
        requests.put(url, headers=headers, json={
            "message": f"Log: {action}", "content": content, "sha": sha
        })
    except Exception:
        pass


# ── Claude extraction ──────────────────────────────────────────
def extract_invoice(file_bytes: bytes, filename: str) -> dict:
    ext = filename.rsplit(".", 1)[-1].lower()
    b64 = base64.standard_b64encode(file_bytes).decode()

    if ext == "pdf":
        file_content = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }
    else:
        media_map = {
            "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "png": "image/png",  "webp": "image/webp",
        }
        file_content = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_map.get(ext, "image/jpeg"),
                "data": b64,
            },
        }

    client = anthropic.Anthropic(api_key=st.secrets["ANTHROPIC_API_KEY"])
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [file_content, {"type": "text", "text": EXTRACT_PROMPT}],
        }],
    )

    text = response.content[0].text.strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    data  = json.loads(match.group() if match else text)
    data["source_file"] = filename
    return data


def save_config(new_config: dict):
    """Save config to local disk and push to GitHub."""
    try:
        with open("config.yaml", "w", encoding="utf-8") as f:
            yaml.dump(new_config, f, allow_unicode=True, default_flow_style=False)
    except Exception:
        pass
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO", "gil-hue/invoice-extractor")
        if not token:
            return
        url     = f"https://api.github.com/repos/{repo}/contents/config.yaml"
        headers = {"Authorization": f"token {token}"}
        r       = requests.get(url, headers=headers)
        if r.status_code != 200:
            return
        sha     = r.json()["sha"]
        content = base64.b64encode(
            yaml.dump(new_config, allow_unicode=True, default_flow_style=False).encode()
        ).decode()
        requests.put(url, headers=headers, json={
            "message": "Update config (user management)",
            "content": content,
            "sha": sha,
        })
    except Exception:
        pass


def safe_extract(file_bytes: bytes, filename: str) -> dict:
    try:
        return extract_invoice(file_bytes, filename)
    except Exception as e:
        empty = {k: "" for k in COL_ORDER}
        empty["source_file"] = filename
        empty["notes"] = f"שגיאה: {str(e)}"
        return empty


# ── Excel builder ──────────────────────────────────────────────
def _sum_col(df: pd.DataFrame, col: str) -> float:
    try:
        return float(pd.to_numeric(df[col], errors="coerce").sum())
    except Exception:
        return 0.0


def build_invoice_excel(df: pd.DataFrame) -> bytes:
    wb        = Workbook()
    dark_fill = PatternFill("solid", start_color="1A4731")
    mid_fill  = PatternFill("solid", start_color="2E7D52")
    light     = PatternFill("solid", start_color="D4EDDE")
    white     = PatternFill("solid", start_color="FFFFFF")
    alt       = PatternFill("solid", start_color="EBF7F0")
    title_f   = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    hdr_f     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_f    = Font(name="Arial", size=11)
    total_f   = Font(name="Arial", bold=True, size=11, color="1A4731")
    thin      = Side(style="thin", color="BFBFBF")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")

    # ── Sheet 1: Summary ────────────────────────────────────────
    ws_s = wb.active
    ws_s.title = "סיכום"
    ws_s.merge_cells("A1:C1")
    ws_s["A1"] = "סיכום חשבוניות — " + date.today().strftime("%d/%m/%Y")
    ws_s["A1"].font = title_f
    ws_s["A1"].fill = dark_fill
    ws_s["A1"].alignment = center
    ws_s.row_dimensions[1].height = 32

    summary_rows = [
        ("מספר חשבוניות",       len(df)),
        ('סה"כ לפני מע"מ (₪)', _sum_col(df, "amount_before_vat")),
        ('סה"כ מע"מ (₪)',       _sum_col(df, "vat_amount")),
        ('סה"כ כולל מע"מ (₪)', _sum_col(df, "total_amount")),
    ]
    for i, (label, val) in enumerate(summary_rows):
        r    = i + 2
        fill = light if i % 2 == 0 else white
        c1   = ws_s.cell(r, 1, label)
        c2   = ws_s.cell(r, 3, val if isinstance(val, int) else round(val, 2))
        for c in [c1, c2]:
            c.font   = total_f
            c.fill   = fill
            c.border = brd
        c1.alignment = left
        c2.alignment = center
        ws_s.cell(r, 2).fill   = fill
        ws_s.cell(r, 2).border = brd
        ws_s.row_dimensions[r].height = 22
    ws_s.column_dimensions["A"].width = 32
    ws_s.column_dimensions["B"].width = 10
    ws_s.column_dimensions["C"].width = 22

    # ── Sheet 2: Invoice data ────────────────────────────────────
    ws_d  = wb.create_sheet("פרטי חשבוניות")
    cols  = [c for c in COL_ORDER if c in df.columns]
    widths = {
        "source_file": 28, "invoice_number": 18, "date": 14,
        "supplier_name": 26, "supplier_vat": 16, "customer_name": 26,
        "customer_vat": 16, "document_type": 22, "payment_method": 18,
        "amount_before_vat": 18, "vat_amount": 14, "total_amount": 16,
        "notes": 40,
    }
    for ci, col in enumerate(cols, 1):
        c = ws_d.cell(1, ci, COL_LABELS.get(col, col))
        c.font = hdr_f; c.fill = mid_fill; c.alignment = center; c.border = brd
        ws_d.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)
    ws_d.row_dimensions[1].height = 26
    ws_d.freeze_panes = "A2"

    for ri, row in enumerate(df[cols].itertuples(index=False), 2):
        fill = white if ri % 2 == 0 else alt
        for ci, val in enumerate(row, 1):
            c = ws_d.cell(ri, ci, str(val) if pd.notna(val) else "")
            c.font = data_f; c.fill = fill; c.border = brd; c.alignment = left
        ws_d.row_dimensions[ri].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════
auth_status = st.session_state.get("authentication_status")

if auth_status is not True:
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown('<div class="login-title">🧾 חשבוניות חכמות</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-sub">כניסה למערכת</div>', unsafe_allow_html=True)
        authenticator.login(location="main")
        if auth_status is False:
            st.error("❌ שם משתמש או סיסמה שגויים")
    st.stop()

# ── Logout bar ─────────────────────────────────────────────────
col_u, col_lo = st.columns([9, 1])
with col_lo:
    authenticator.logout("התנתק")

current_user = st.session_state.get("username", "")
is_admin     = config["credentials"]["usernames"].get(current_user, {}).get("role") == "admin"

st.markdown("""
<div class="title-bar">
    <h1>🧾 חשבוניות חכמות</h1>
    <p>העלה חשבוניות סרוקות או PDF — חלץ נתונים אוטומטית עם Claude AI</p>
</div>
""", unsafe_allow_html=True)

# ── Tabs ───────────────────────────────────────────────────────
if is_admin:
    tab_main, tab_log, tab_users = st.tabs(["🧾 חילוץ חשבוניות", "📋 לוג פעולות", "👥 ניהול משתמשים"])
else:
    tab_main  = st.tabs(["🧾 חילוץ חשבוניות"])[0]
    tab_log   = None
    tab_users = None

# ══════════════════════════════════════════════════════════════
#  LOG TAB (admin only)
# ══════════════════════════════════════════════════════════════
if is_admin and tab_users:
    with tab_users:
        st.markdown('<div class="section-header">👥 ניהול משתמשים</div>', unsafe_allow_html=True)

        users = config["credentials"]["usernames"]

        # ── Users table ────────────────────────────────────────
        user_rows = [
            {
                "👤 שם משתמש": uname,
                "📛 שם מלא":   uinfo.get("name", ""),
                "📧 אימייל":   uinfo.get("email", ""),
                "🔑 תפקיד":    "אדמין" if uinfo.get("role") == "admin" else "משתמש",
            }
            for uname, uinfo in users.items()
        ]
        st.dataframe(pd.DataFrame(user_rows), use_container_width=True, hide_index=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Add user ────────────────────────────────────────────
        with st.expander("➕ הוסף משתמש חדש", expanded=False):
            with st.form("add_user_form"):
                c1, c2 = st.columns(2)
                with c1:
                    nu_username = st.text_input("שם משתמש *")
                    nu_name     = st.text_input("שם מלא")
                with c2:
                    nu_email    = st.text_input("אימייל")
                    nu_password = st.text_input("סיסמה *", type="password")
                nu_role = st.selectbox("תפקיד", ["משתמש", "אדמין"])
                if st.form_submit_button("✅ הוסף משתמש", use_container_width=True):
                    if not nu_username or not nu_password:
                        st.error("שם משתמש וסיסמה הם שדות חובה")
                    elif nu_username in users:
                        st.error(f"שם המשתמש '{nu_username}' כבר קיים")
                    else:
                        hashed = bcrypt.hashpw(nu_password.encode(), bcrypt.gensalt()).decode()
                        config["credentials"]["usernames"][nu_username] = {
                            "name":     nu_name,
                            "email":    nu_email,
                            "password": hashed,
                            "role":     "admin" if nu_role == "אדמין" else "user",
                        }
                        save_config(config)
                        write_log("ניהול משתמשים", f"נוסף משתמש: {nu_username}")
                        st.success(f"✅ משתמש '{nu_username}' נוסף בהצלחה!")
                        st.rerun()

        # ── Change password ─────────────────────────────────────
        with st.expander("🔒 שינוי סיסמה למשתמש", expanded=False):
            with st.form("change_pass_form"):
                cp_user = st.selectbox("בחר משתמש", list(users.keys()), key="cp_user")
                cp_pass = st.text_input("סיסמה חדשה *", type="password")
                if st.form_submit_button("✅ עדכן סיסמה", use_container_width=True):
                    if not cp_pass:
                        st.error("יש להזין סיסמה חדשה")
                    else:
                        hashed = bcrypt.hashpw(cp_pass.encode(), bcrypt.gensalt()).decode()
                        config["credentials"]["usernames"][cp_user]["password"] = hashed
                        save_config(config)
                        write_log("ניהול משתמשים", f"עודכנה סיסמה: {cp_user}")
                        st.success(f"✅ סיסמת '{cp_user}' עודכנה בהצלחה!")

        # ── Delete user ─────────────────────────────────────────
        with st.expander("🗑️ מחיקת משתמש", expanded=False):
            deletable = [u for u in users.keys() if u != current_user]
            if deletable:
                with st.form("del_user_form"):
                    del_user = st.selectbox("בחר משתמש למחיקה", deletable, key="del_user")
                    st.warning(f"⚠️ פעולה זו תמחק לצמיתות את המשתמש '{del_user}'")
                    if st.form_submit_button("🗑️ מחק משתמש", use_container_width=True):
                        del config["credentials"]["usernames"][del_user]
                        save_config(config)
                        write_log("ניהול משתמשים", f"נמחק משתמש: {del_user}")
                        st.success(f"✅ משתמש '{del_user}' נמחק בהצלחה!")
                        st.rerun()
            else:
                st.info("אין משתמשים נוספים למחיקה (לא ניתן למחוק את עצמך)")


if is_admin and tab_log:
    with tab_log:
        st.markdown('<div class="section-header">📋 לוג פעולות</div>', unsafe_allow_html=True)
        col_r, col_f = st.columns([1, 3])
        with col_r:
            if st.button("🔄 רענן", key="refresh_log"):
                st.rerun()
        with col_f:
            filter_action = st.selectbox(
                "סנן לפי פעולה",
                ["הכל", "העלאה וחילוץ", "הורדת Excel"],
                key="log_filter",
            )

        logs = read_logs()
        if logs:
            if filter_action != "הכל":
                logs = [l for l in logs if l.get("action") == filter_action]
            if logs:
                action_icons = {"העלאה וחילוץ": "📂", "הורדת Excel": "⬇️"}
                df_log = pd.DataFrame([{
                    "🕐 תאריך ושעה": l.get("timestamp", ""),
                    "👤 משתמש":       l.get("user", ""),
                    "פעולה":          action_icons.get(l.get("action", ""), "•") + " " + l.get("action", ""),
                    "📝 פרטים":       l.get("details", ""),
                } for l in logs])

                ITEMS_PER_PAGE = 20
                total_pages    = max(1, (len(df_log) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
                if st.session_state.get("_last_log_filter") != filter_action:
                    st.session_state["log_page"]         = 0
                    st.session_state["_last_log_filter"] = filter_action
                if "log_page" not in st.session_state:
                    st.session_state["log_page"] = 0
                page = min(st.session_state["log_page"], total_pages - 1)
                st.session_state["log_page"] = page
                start   = page * ITEMS_PER_PAGE
                df_page = df_log.iloc[start : start + ITEMS_PER_PAGE]

                st.dataframe(df_page, use_container_width=True, hide_index=True)
                st.caption(f'סה"כ: {len(df_log)} פעולות | עמוד {page + 1} מתוך {total_pages}')

                if total_pages > 1:
                    half       = 2
                    p_start    = max(0, page - half)
                    p_end      = min(total_pages, p_start + 5)
                    if p_end - p_start < 5:
                        p_start = max(0, p_end - 5)
                    page_range = list(range(p_start, p_end))
                    pcols = st.columns([0.5] + [0.3] * len(page_range) + [0.5], gap="small")
                    with pcols[0]:
                        if st.button("‹", key="pg_prev", disabled=(page == 0), type="secondary"):
                            st.session_state["log_page"] = page - 1
                            st.rerun()
                    for i, pg in enumerate(page_range):
                        with pcols[i + 1]:
                            b_type = "primary" if pg == page else "secondary"
                            if st.button(str(pg + 1), key=f"pg_{pg}", type=b_type):
                                st.session_state["log_page"] = pg
                                st.rerun()
                    with pcols[-1]:
                        if st.button("›", key="pg_next", disabled=(page == total_pages - 1), type="secondary"):
                            st.session_state["log_page"] = page + 1
                            st.rerun()

                st.markdown("---")
                buf = io.BytesIO()
                df_log.to_excel(buf, index=False)
                buf.seek(0)
                st.download_button(
                    "⬇️ ייצא לוג ל-Excel",
                    data=buf.read(),
                    file_name=f"InvoiceLog_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.info("אין פעולות מסוג זה בלוג")
        else:
            st.info("הלוג ריק — פעולות יופיעו כאן לאחר שימוש באפליקציה")
            st.caption("💡 הלוג מצריך הגדרת GITHUB_TOKEN ב-Secrets לשמירה קבועה")

# ══════════════════════════════════════════════════════════════
#  MAIN TAB
# ══════════════════════════════════════════════════════════════
with tab_main:

    # ── Step 1: Upload ─────────────────────────────────────────
    st.markdown('<div class="section-header">📂 שלב 1 — העלאת חשבוניות</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "גרור חשבוניות לכאן (PDF, JPG, PNG, WEBP)",
        type=["pdf", "jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded:
        new_names = sorted([f.name for f in uploaded])
        if new_names != st.session_state.get("_last_uploaded_names", []):
            st.session_state["_last_uploaded_names"] = new_names
            st.session_state["uploaded_files"]       = {f.name: f.read() for f in uploaded}
            st.session_state.pop("extracted_df", None)

    uploaded_files: dict = st.session_state.get("uploaded_files", {})

    if not uploaded_files:
        st.info("⬆️ העלה חשבונית אחת או יותר (PDF / תמונה סרוקה) כדי להמשיך.")
    else:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                f'<div class="metric-card"><div class="num">{len(uploaded_files)}</div>'
                f'<div class="lbl">קבצים הועלו</div></div>',
                unsafe_allow_html=True,
            )
        with c2:
            ext_count = len(st.session_state.get("extracted_df", pd.DataFrame()))
            st.markdown(
                f'<div class="metric-card"><div class="num">{ext_count}</div>'
                f'<div class="lbl">חשבוניות חולצו</div></div>',
                unsafe_allow_html=True,
            )
        st.markdown("<br>", unsafe_allow_html=True)

        with st.expander("📋 קבצים שהועלו", expanded=True):
            hdr1, hdr2, hdr3, hdr4 = st.columns([0.4, 4, 1.5, 1])
            hdr1.markdown("**🗑️**")
            hdr2.markdown("**📄 שם קובץ**")
            hdr3.markdown("**📏 גודל**")
            hdr4.markdown("**📌 סוג**")
            st.divider()
            to_remove = []
            for name, b in uploaded_files.items():
                col_cb, col_name, col_size, col_type = st.columns([0.4, 4, 1.5, 1])
                with col_cb:
                    if st.checkbox("", key=f"_rm_{name}", label_visibility="hidden"):
                        to_remove.append(name)
                col_name.write(name)
                col_size.write(f"{len(b)/1024:.1f} KB")
                col_type.write(name.rsplit('.', 1)[-1].upper())
            if to_remove:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button(f"🗑️ הסר {len(to_remove)} קבצים נבחרים", type="secondary", key="remove_files"):
                    for fname in to_remove:
                        st.session_state["uploaded_files"].pop(fname, None)
                        st.session_state.pop(f"_rm_{fname}", None)
                    # ⚠️ אל תעדכן _last_uploaded_names — ה-file_uploader עדיין מחזיר את כל הקבצים
                    # אם נעדכן לרשימה קצרה, הריראן הבא יזהה מיסמאץ' וישחזר הכל
                    st.session_state.pop("extracted_df", None)
                    st.rerun()

        # ── Step 2: Extract ────────────────────────────────────
        st.markdown(
            '<div class="section-header">🤖 שלב 2 — חילוץ נתונים עם Claude AI</div>',
            unsafe_allow_html=True,
        )

        if st.button("🚀 חלץ נתונים מהחשבוניות", use_container_width=True):
            results  = []
            progress = st.progress(0, text="מכין...")
            total    = len(uploaded_files)
            for i, (fname, fbytes) in enumerate(uploaded_files.items()):
                progress.progress(i / total, text=f"מחלץ: {fname}...")
                results.append(safe_extract(fbytes, fname))
            progress.progress(1.0, text="✅ הושלם!")

            df = pd.DataFrame([{col: r.get(col, "") for col in COL_ORDER} for r in results])
            st.session_state["extracted_df"] = df
            write_log(
                "העלאה וחילוץ",
                f"{len(results)} חשבוניות: {', '.join(uploaded_files.keys())[:150]}",
            )

        # ── Step 3: Edit & download ────────────────────────────
        if "extracted_df" in st.session_state:
            df_base: pd.DataFrame = st.session_state["extracted_df"]

            st.markdown(
                '<div class="section-header">✏️ שלב 3 — עריכה ואישור</div>',
                unsafe_allow_html=True,
            )
            st.caption("ניתן לערוך ישירות בטבלה לפני ההורדה")

            display_rename = {c: COL_LABELS.get(c, c) for c in COL_ORDER if c in df_base.columns}
            df_display     = df_base.rename(columns=display_rename)

            # ── Per-row delete checkboxes (stable keys) ─────────
            rows_to_delete = [
                i for i in df_display.index
                if st.session_state.get(f"_del_row_{i}")
            ]
            if rows_to_delete:
                if st.button(f"🗑️ הסר {len(rows_to_delete)} חשבוניות נבחרות", type="secondary", key="del_rows"):
                    new_df = st.session_state["extracted_df"].drop(index=rows_to_delete).reset_index(drop=True)
                    for i in rows_to_delete:
                        st.session_state.pop(f"_del_row_{i}", None)
                    st.session_state["extracted_df"] = new_df
                    st.rerun()

            df_display.insert(0, "🗑️", [st.session_state.get(f"_del_row_{i}", False) for i in df_display.index])

            edited_df = st.data_editor(
                df_display,
                use_container_width=True,
                hide_index=True,
                key="invoice_editor",
                column_config={
                    "🗑️":                            st.column_config.CheckboxColumn("🗑️ הסר", width="small"),
                    COL_LABELS["source_file"]:       st.column_config.TextColumn(disabled=True, width="medium"),
                    COL_LABELS["invoice_number"]:    st.column_config.TextColumn(width="small"),
                    COL_LABELS["date"]:              st.column_config.TextColumn(width="small"),
                    COL_LABELS["amount_before_vat"]: st.column_config.TextColumn(width="small"),
                    COL_LABELS["vat_amount"]:        st.column_config.TextColumn(width="small"),
                    COL_LABELS["total_amount"]:      st.column_config.TextColumn(width="small"),
                    COL_LABELS["notes"]:             st.column_config.TextColumn(width="large"),
                },
            )

            # sync checkbox column back to session_state
            for i in df_display.index:
                st.session_state[f"_del_row_{i}"] = bool(edited_df.at[i, "🗑️"])

            edited_df = edited_df.drop(columns=["🗑️"])

            # ── Totals strip ───────────────────────────────────
            st.markdown("<br>", unsafe_allow_html=True)
            try:
                total_before = pd.to_numeric(edited_df[COL_LABELS["amount_before_vat"]], errors="coerce").sum()
                total_vat    = pd.to_numeric(edited_df[COL_LABELS["vat_amount"]],        errors="coerce").sum()
                total_all    = pd.to_numeric(edited_df[COL_LABELS["total_amount"]],      errors="coerce").sum()
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown(
                        f'<div class="metric-card"><div class="num">₪{total_before:,.2f}</div>'
                        f'<div class="lbl">סה"כ לפני מע"מ</div></div>',
                        unsafe_allow_html=True,
                    )
                with c2:
                    st.markdown(
                        f'<div class="metric-card"><div class="num">₪{total_vat:,.2f}</div>'
                        f'<div class="lbl">סה"כ מע"מ</div></div>',
                        unsafe_allow_html=True,
                    )
                with c3:
                    st.markdown(
                        f'<div class="metric-card"><div class="num">₪{total_all:,.2f}</div>'
                        f'<div class="lbl">סה"כ כולל מע"מ</div></div>',
                        unsafe_allow_html=True,
                    )
            except Exception:
                pass

            st.markdown("<br>", unsafe_allow_html=True)

            # Build Excel from the current (possibly edited) data
            reverse_labels = {v: k for k, v in COL_LABELS.items()}
            df_for_excel   = edited_df.rename(columns=reverse_labels)
            out_name       = f"Invoices_{date.today().strftime('%Y-%m-%d')}.xlsx"
            excel_bytes    = build_invoice_excel(df_for_excel)

            st.markdown(
                f'<div class="success-box">✅ {len(df_base)} חשבוניות מוכנות להורדה</div>',
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            if st.download_button(
                label="⬇️ הורד Excel עם כל הנתונים",
                data=excel_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            ):
                write_log("הורדת Excel", f"{out_name} | {len(df_base)} חשבוניות")
